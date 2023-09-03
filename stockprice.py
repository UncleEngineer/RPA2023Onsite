from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time

from openpyxl import load_workbook

# load file
loadfile = 'stockprice.xlsx'
myworkbook = load_workbook(loadfile)
mysheet = myworkbook.active
allrow = len(mysheet['A']) + 1

stocklist = []
pricelist = []

for i in range(2, allrow):
    stocklist.append(mysheet.cell(row=i, column=1).value)

# print(stocklist)

# ระบุตำแหน่งของ Chromedriver
path = r'C:\Users\User\OneDrive\Desktop\Basic Python\chromedriver.exe'
service = Service(path)

# สั่งเปิด Chrome เต็มหน้าจอ
option = webdriver.ChromeOptions()
option.add_argument('--start-maximized')

url = 'https://www.google.com/'
driver = webdriver.Chrome(service=service, options=option)

for s in stocklist:
    driver.get(url)

    stock = driver.find_element(By.NAME, "q")
    stock.send_keys('ราคาหุ้น ' + s)
    stock.send_keys(Keys.RETURN)

    xpath = '//*[@id="knowledge-finance-wholepage__entity-summary"]/div[3]/g-card-section/div/g-card-section/div[2]/div[1]/span[1]/span/span[1]'
    price = driver.find_element(By.XPATH, xpath)
    pricelist.append(float(price.text))

    for r, p in enumerate(pricelist, 2):
        writedata = mysheet.cell(row=r, column=2)
        writedata.value = p

    time.sleep(2)
    print(s, p)

driver.quit()
myworkbook.save(loadfile)