from openpyxl import load_workbook

# load file
loadfile = 'stockprice.xlsx'
myworkbook = load_workbook(loadfile)
mysheet = myworkbook.active
allrow = len(mysheet['A']) + 1

stocklist = []

for i in range(2, allrow):
    stocklist.append(mysheet.cell(row=i, column=1).value)

print(stocklist)